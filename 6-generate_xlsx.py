import os
import glob
import pandas as pd
from openpyxl.styles import Alignment, Font

def parse_rnk_file_to_dataframe(filepath):
    column_names = [
        'Mol No', 'Score', 'S(PLP)', 'S(hbond)', 'S(cho)', 'S(metal)',
        'DE(clash)', 'DE(tors)', 'intcor', 'time'
    ]
    try:
        df = pd.read_csv(filepath, skiprows=4, sep=r'\s+', names=column_names)
        df['Mol No'] = pd.to_numeric(df['Mol No'], errors='coerce')
        df.dropna(subset=['Mol No'], inplace=True)
        df['Mol No'] = df['Mol No'].astype(int)
        df.drop_duplicates(subset='Mol No', keep='first', inplace=True)
        df.set_index('Mol No', inplace=True)
        return df
    except FileNotFoundError:
        print(f"WARNING: File {filepath} was not found and will be skipped.")
        return None
    except Exception as e:
        print(f"An error occurred while reading {filepath}: {e}")
        return None

def main():
    monomers_path = "../monomers"

    if not os.path.isdir(monomers_path):
        print(f"ERROR: The monomers folder was not found in: '{monomers_path}'")
        return

    search_pattern = os.path.join(monomers_path, "Variant*_*_monomer")
    variant_folders = sorted(glob.glob(search_pattern))

    if not variant_folders:
        print(f"ERROR: No 'Variant...' folders were found inside '{monomers_path}'")
        return

    print(f"Found {len(variant_folders)} variant folders to process...")

    df_agg = pd.DataFrame({'Score': list(range(1, 200))})
    df_agg.set_index('Score', inplace=True)

    for variant_path in variant_folders:
        folder_name = os.path.basename(variant_path)
        print(f"Processing: {folder_name}")

        path_asn = os.path.join(variant_path, 'L-Asn_m1', 'L-Asn_m1.rnk')
        path_gln = os.path.join(variant_path, 'L-Gln_m1', 'L-Gln_m1.rnk')

        df_asn = parse_rnk_file_to_dataframe(path_asn)
        df_gln = parse_rnk_file_to_dataframe(path_gln)

        score = list(range(1, 201))
        scores_asn = df_asn.reindex(score)['Score'] if df_asn is not None else pd.Series(index=score, dtype=float)
        scores_gln = df_gln.reindex(score)['Score'] if df_gln is not None else pd.Series(index=score, dtype=float)

        df_agg[f'{folder_name}_Asn'] = scores_asn.values
        df_agg[f'{folder_name}_Gln'] = scores_gln.values
        df_agg[f'spacer_{folder_name}'] = ''

    output_filename = 'results.xlsx'
    print(f"\nCreating Excel file: {output_filename}")

    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df_agg.reset_index().to_excel(writer, sheet_name='Results', startrow=2, header=False, index=False)

        workbook = writer.book
        worksheet = writer.sheets['Results']

        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        cell_pose = worksheet['A1']
        cell_pose.value = "Pose"
        cell_pose.font = header_font
        cell_pose.alignment = center_align
        worksheet.merge_cells('A1:A2')

        for i, variant_path in enumerate(variant_folders):
            folder_name = os.path.basename(variant_path)
            start_col = 2 + (i * 3)
            end_col = start_col + 1

            cell_header = worksheet.cell(row=1, column=start_col)
            cell_header.value = folder_name
            cell_header.font = header_font
            cell_header.alignment = center_align
            worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            worksheet.cell(row=2, column=start_col).value = "Asparagine"
            worksheet.cell(row=2, column=end_col).value = "Glutamine"

    print("\nProcess completed successfully!")
    print(f"File '{output_filename}' created at: {os.path.abspath(output_filename)}")

if __name__ == "__main__":
    main()
